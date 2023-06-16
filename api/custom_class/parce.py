import logging

import openpyxl

from Book_backend import celery_app as app
from api.models import Author, Artworks, Genre
import time


class ParseXML:
    """
    Класс для парсинга книг
    """

    def __init__(self, file_path: str):
        self.file_path = file_path

    @staticmethod
    def create_author(fio):
        """
        Создает автора и возвращает объект Author
        """
        author, _ = Author.objects.get_or_create(name=fio)
        return author

    @staticmethod
    def create_genres(zhanr):
        """
        Создает жанры и возвращает список их идентификаторов
        """
        genre_objs = []
        zhanr = zhanr.split(',')
        for genre in zhanr:
            obj, _ = Genre.objects.get_or_create(name=genre)
            genre_objs.append(obj.id)
        return genre_objs

    @staticmethod
    def create_artwork(proizvedenie, god, nazvanie_fayla):
        """
        Создает произведение и возвращает объект Artworks
        """
        return Artworks.objects.get_or_create(
            name=proizvedenie,
            defaults={
                'date': god,
                'file': f'/media/book/{nazvanie_fayla}.epub',
            }
        )


    def parse_excel_file(self):
        """
        Функция lzk парсинга книг
        :return:
        """
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):

            fio, proizvedenie, nazvanie_fayla, god, forma, zhanr, tagi = row[:7]
            if Artworks.objects.filter(name=proizvedenie).exists():
                continue
            start_time = time.time()
            author = self.create_author(fio)
            genres = self.create_genres(zhanr=zhanr)
            artworks, _ = self.create_artwork(
                proizvedenie=proizvedenie,
                god=god,
                nazvanie_fayla=nazvanie_fayla,
            )
            artworks.author.add(author)
            artworks.genres.set(genres)
            end_time = time.time()
            elapsed_time = end_time - start_time
            logging.error(f"Elapsed time for parse_excel_file: {elapsed_time} seconds")


    def create_genres_func(self):
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            zhanr = row[5]
            self.create_genres(zhanr=zhanr)

    def create_author_func(self):
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            fio, proizvedenie, nazvanie_fayla, god, forma, zhanr, tagi = row[:7]
            if Artworks.objects.filter(name=proizvedenie).exists():
                continue
            logging.error('start')
            self.create_author(fio)